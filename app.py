import streamlit as st
import pandas as pd
import numpy as np
import io
import re
import plotly.graph_objects as fgo
from openpyxl.chart import BarChart, LineChart, Reference

# 设置网页标题与布局
st.set_page_config(page_title="语音数据智能分析工具", layout="wide")

st.title("📊 南京金鹰世界G酒店-语音运营数据分析")
st.markdown("页面与Excel已同步支持 1:1 复刻双轴复合图表（三折线+单柱）。")

# ----------------- 🛠️ 数据矩阵上传区 -----------------
st.markdown("### 📥 数据矩阵上传中心")
col_up1, col_up2, col_up3 = st.columns(3)

with col_up1:
    st.markdown("**▼ 基础配置表**")
    uploaded_ext_file = st.file_uploader("分机号参考表 (CSV/XLSX)", type=["csv", "xlsx"], key="ext")

with col_up2:
    st.markdown("**▼ 空间 A：上线后运营时段**")
    uploaded_post_file = st.file_uploader("上线后-云总机通话详单", type=["csv", "xlsx"], key="post")

with col_up3:
    st.markdown("**▼ 空间 B：上线前历史时段**")
    uploaded_pre_file = st.file_uploader("上线前-历史通话详单 (选填)", type=["csv", "xlsx"], key="pre")


def clean_to_int_str(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    return val_str


def extract_hour_from_datetime(val):
    """从标准呼叫时间字符串中精准切出小时"""
    if pd.isna(val):
        return None
    val_str = str(val).strip()
    match = re.search(r'(\d{1,2}):\d{2}:\d{2}', val_str)
    if match:
        try:
            h = int(match.group(1))
            if 0 <= h <= 23:
                return h
        except:
            pass
    return None


# 统一的数据清洗计算引擎
def process_data(df_raw, df_ext):
    df_call = df_raw.copy()
    df_call.columns = df_call.columns.astype(str).str.strip()
    
    required_cols = ['主叫号码', '通话状态', 'AI通话状态', '人工通话状态', '呼叫时间']
    if any(col not in df_call.columns for col in required_cols):
        return None
        
    df_call['主叫号码_clean'] = df_call['主叫号码'].apply(clean_to_int_str)
    valid_extensions = set(df_ext['分机号_clean'].dropna().unique())
    
    # 1. 判定客房接入
    df_call['房间是否接入'] = df_call['主叫号码_clean'].apply(
        lambda x: '客房' if (x and x in valid_extensions) else '--'
    )
    
    # 2. 从“呼叫时间”直接解析出标准小时
    df_call['⚡标准时段'] = df_call['呼叫时间'].apply(extract_hour_from_datetime)

    # 3. 判定最终成功接通
    def check_final_success(row):
        if row['房间是否接入'] != '客房':  # 已经完美修复为简体字 “间”
            return '--'
        m = str(row['通话状态']).strip()
        n = str(row['AI通话状态']).strip()
        o = str(row['人工通话状态']).strip()
        
        if m == '接通' or n == '接通' or o == '接通':
            return '是'
        return '否'
    df_call['最终成功接通'] = df_call.apply(check_final_success, axis=1)
    
    # 4. 判定接通方式划分
    def check_connect_type(row):
        if row['房间是否接入'] != '客房': 
            return '--'
        m = str(row['通话状态']).strip()
        n = str(row['AI通话状态']).strip()
        o = str(row['人工通话状态']).strip()
        
        if m == '接通' and n == '接通' and o == '接通':
            return '进入AI后，再转接人工，且人工接通'
        elif m == '接通' and n == '接通' and o == '未接通':
            return 'AI接通，转接人工，人工未接通'
        elif m == '接通' and n == '接通' and (o == '--' or o == ''):
            return '进入AI后，AI直接完成，未转接人工'
        elif m == '接通' and (n == '--' or n == '') and o == '接通':
            return '直接进入人工，且人工接通'
        elif m == '未接通' and (n == '--' or n == '') and (o == '--' or o == ''):
            return '客人主动挂断'
        elif m == '未接通' and (n == '--' or n == '') and o == '未接通':
            return '直接进入人工且最终未接通'
        elif m == '接通' and (n == '--' or n == '') and (o == '--' or o == ''):
            return '直接进入人工，且人工接通'
        else:
            return '其他未划分场景'
            
    df_call['接通方式'] = df_call.apply(check_connect_type, axis=1)
    return df_call


# ----------------- 📈 看板数据渲染 -----------------
if uploaded_post_file and uploaded_ext_file:
    with st.spinner("双时段效益及24H时段多维对账中..."):
        try:
            if uploaded_ext_file.name.endswith('.csv'):
                df_ext = pd.read_csv(uploaded_ext_file)
            else:
                df_ext = pd.read_excel(uploaded_ext_file)
            df_ext.columns = df_ext.columns.astype(str).str.strip()
            df_ext['分机号_clean'] = df_ext['分机号'].apply(clean_to_int_str)
            
            if uploaded_post_file.name.endswith('.csv'):
                df_post_raw = pd.read_csv(uploaded_post_file)
            else:
                df_post_raw = pd.read_excel(uploaded_post_file)
                
            st.session_state['df_post_res'] = process_data(df_post_raw, df_ext)
            
            if st.session_state['df_post_res'] is not None:
                df_rooms_post = st.session_state['df_post_res'][st.session_state['df_post_res']['房间是否接入'] == '客房']
                total_calls = len(df_rooms_post)
                
                # AI接通率大盘
                df_rooms_post['AI通话状态_clean'] = df_rooms_post['AI通话状态'].astype(str).str.strip()
                ai_success = int((df_rooms_post['AI通话状态_clean'] == '接通').sum())
                ai_total = int(df_rooms_post['AI通话状态_clean'].isin(['接通', '未接通']).sum())
                ai_rate_str = f"{ai_success / ai_total:.1%}" if ai_total > 0 else "--"
                is_ai_abnormal = (ai_success != ai_total) if ai_total > 0 else False
                
                # 上线后人工接通率大盘
                connect_types = df_rooms_post['接通方式'].str.strip()
                man_success_post = int(connect_types.isin(['进入AI后，再转接人工，且人工接通', '直接进入人工，且人工接通']).sum())
                man_total_post = int(connect_types.isin(['进入AI后，再转接人工，且人工接通', '直接进入人工，且人工接通', 'AI接通，转接人工，人工未接通', '直接进入人工且最终未接通']).sum())
                online_man_rate_str = f"{man_success_post / man_total_post:.1%}" if man_total_post > 0 else "--"
                
                # 上线后整体接通率大盘
                final_success_count = int((df_rooms_post['最终成功接通'] == '是').sum())
                overall_post_rate = final_success_count / total_calls if total_calls > 0 else 0.0
                overall_post_rate_str = f"{overall_post_rate:.1%}" if total_calls > 0 else "--"
                
                pre_man_rate_str = "--"
                lift_rate_str = "--"
                saved_calls_str = "--"
                
                # 上线后分时放量
                df_h_post = df_rooms_post[df_rooms_post['⚡标准时段'].notna()]
                post_h_total = df_h_post.groupby('⚡标准时段').size()
                post_h_success = df_h_post[df_h_post['最终成功接通'] == '是'].groupby('⚡标准时段').size()
                post_h_man_success = df_h_post[df_h_post['接通方式'].str.strip().isin(['进入AI后，再转接人工，且人工接通', '直接进入人工，且人工接通'])].groupby('⚡标准时段').size()
                post_h_man_total = df_h_post[df_h_post['接通方式'].str.strip().isin(['进入AI后，再转接人工，且人工接通', '直接进入人工，且人工接通', 'AI接通，转接人工，人工未接通', '直接进入人工且最终未接通'])].groupby('⚡标准时段').size()

                pre_h_total, pre_h_success = {}, {}
                if uploaded_pre_file:
                    if uploaded_pre_file.name.endswith('.csv'):
                        df_pre_raw = pd.read_csv(uploaded_pre_file)
                    else:
                        df_pre_raw = pd.read_excel(uploaded_pre_file)
                        
                    st.session_state['df_pre_res'] = process_data(df_pre_raw, df_ext)
                    if st.session_state['df_pre_res'] is not None:
                        df_rooms_pre = st.session_state['df_pre_res'][st.session_state['df_pre_res']['房间是否接入'] == '客房']
                        
                        pre_success = int((df_rooms_pre['最终成功接通'] == '是').sum())
                        pre_total = len(df_rooms_pre)
                        
                        if pre_total > 0:
                            overall_pre_rate = pre_success / pre_total
                            pre_man_rate_str = f"{overall_pre_rate:.1%}"
                            lift_value = overall_post_rate - overall_pre_rate
                            lift_rate_str = f"{lift_value:.1%}"
                            saved_calls_str = f"{int(round(total_calls * lift_value))}.0"
                        
                        # 上线前分时聚合
                        df_h_pre = df_rooms_pre[df_rooms_pre['⚡标准时段'].notna()]
                        pre_h_total = df_h_pre.groupby('⚡标准时段').size()
                        pre_h_success = df_h_pre[df_h_pre['最终成功接通'] == '是'].groupby('⚡标准时段').size()
                
                # 双流合并计算
                web_rows = []
                excel_rows = []
                
                for i in range(24):
                    t_pre = pre_h_total.get(i, 0)
                    s_pre = pre_h_success.get(i, 0)
                    t_post = post_h_total.get(i, 0)
                    s_post = post_h_success.get(i, 0)
                    s_man_post = post_h_man_success.get(i, 0)
                    t_man_post = post_h_man_total.get(i, 0)
                    
                    val_pre_overall = s_pre / t_pre if t_pre > 0 else None
                    val_post_man = s_man_post / t_man_post if t_man_post > 0 else None
                    val_post_overall = s_post / t_post if t_post > 0 else None
                    
                    str_pre_overall = f"{val_pre_overall:.1%}" if val_pre_overall is not None else "--"
                    str_post_man = f"{val_post_man:.1%}" if val_post_man is not None else "--"
                    str_post_overall = f"{val_post_overall:.1%}" if val_post_overall is not None else "--"
                    
                    str_lift = "--"
                    str_saved = "--"
                    val_lift = None
                    val_saved = None
                    
                    if val_pre_overall is not None and val_post_overall is not None:
                        val_lift = val_post_overall - val_pre_overall
                        val_saved = val_lift * t_post
                        str_saved = f"{val_saved:.1f}"
                        if val_lift > 0:
                            str_lift = f"↑ {val_lift:.1%}"
                        elif val_lift < 0:
                            str_lift = f"{val_lift:.1%} ↓"
                        else:
                            str_lift = "0.0%"
                    
                    web_rows.append({
                        "时段": f"{i}:00",
                        "整体接通率（人工）[上线前]": str_pre_overall,
                        "人工接通率[上线后]": str_post_man,
                        "整体接通率（人工+AI）[上线后]": str_post_overall,
                        "整体接通率提升（上线后）": str_lift,
                        "减少电话漏接量（通）": str_saved
                    })
                    
                    excel_rows.append({
                        "时段": f"{i}:00",
                        "整体接通率（人工）[上线前]": val_pre_overall,
                        "人工接通率[上线后]": val_post_man,
                        "整体接通率（人工+AI）[上线后]": val_post_overall,
                        "整体接通率提升（上线后）": val_lift,
                        "减少电话漏接量（通）": val_saved
                    })
                
                df_web_display = pd.DataFrame(web_rows)
                df_excel_clean = pd.DataFrame(excel_rows)
                
                # --- 前端面板渲染 ---
                st.markdown("### 📞 PART1：酒店电话数据")
                st.info(f"**总来电量（所有启用AI的客房呼出的电话量）：{total_calls}**")
                
                st.markdown("#### **电话接通率情况**")
                col_b1, col_b2, col_b3, col_b4, col_b5, col_b6 = st.columns(6)
                with col_b1: st.metric(label="上线前整体接通率（人工）", value=pre_man_rate_str)
                with col_b2: st.metric(label="上线后人工接通率", value=online_man_rate_str)
                with col_b3:
                    if is_ai_abnormal:
                        st.markdown(f"<div style='background-color:#FFD2D2; padding:5px; border-radius:5px; border-left:3px solid #FF0000;'><p style='margin:0; color:#550000; font-size:14px;'>AI接通率</p><h3 style='margin:5px 0 0 0; color:#CC0000;'>{ai_rate_str}</h3></div>", unsafe_allow_html=True)
                    else:
                        st.metric(label="AI接通率", value=ai_rate_str)
                with col_b4: st.metric(label="上线后整体接通率（人工+AI）", value=overall_post_rate_str)
                with col_b5: st.metric(label="整体接通率提升", value=lift_rate_str)
                with col_b6: st.metric(label="减少电话漏接量", value=saved_calls_str)
                
                st.markdown("---")
                st.markdown("#### ⏰ 分时段接通率情况")
                st.dataframe(df_web_display, width='stretch')
                
                # ----------------- 🎨 网页端交互图表绘制 (Plotly) -----------------
                st.markdown("#### 📉 分时段接通率趋势与挽回话务复盘画布")
                fig = fgo.Figure()
                hours_x = df_excel_clean["时段"]
                
                # 次轴：减少电话漏接量（修长背景柱）
                fig.add_trace(fgo.Bar(
                    x=hours_x, y=df_excel_clean["减少电话漏接量（通）"],
                    name="减少电话漏接量（通）", yaxis="y2",
                    marker_color="rgba(173, 216, 230, 0.4)", marker_line_width=0
                ))
                # 主轴：上线前整体接通率（灰色虚线）
                fig.add_trace(fgo.Scatter(
                    x=hours_x, y=df_excel_clean["整体接通率（人工）[上线前]"],
                    name="整体接通率（人工）[上线前]", mode="lines+markers",
                    line=dict(color="darkgray", width=2, dash="dash"), marker=dict(symbol="circle", size=6)
                ))
                # 主轴：上线后人工接通率（浅蓝色实线）
                fig.add_trace(fgo.Scatter(
                    x=hours_x, y=df_excel_clean["人工接通率[上线后]"],
                    name="人工接通率[上线后]", mode="lines+markers",
                    line=dict(color="#636EFA", width=2), marker=dict(symbol="circle", size=6)
                ))
                # 主轴：上线后整体接通率（深蓝色实线）
                fig.add_trace(fgo.Scatter(
                    x=hours_x, y=df_excel_clean["整体接通率（人工+AI）[上线后]"],
                    name="整体接通率（人工+AI）[上线后]", mode="lines+markers",
                    line=dict(color="#19D3BF", width=3), marker=dict(symbol="circle", size=8)
                ))
                
                fig.update_layout(
                    template="plotly_white",
                    xaxis=dict(title="24小时分时时段"),
                    yaxis=dict(title="接通率 (%)", tickformat=".0%", range=[0, 1.1]),
                    yaxis2=dict(title="减少电话漏接量 (通)", overlaying="y", side="right", showgrid=False),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
                )
                st.plotly_chart(fig, width='stretch')
                
                # ----------------- 🛠️ Excel 原生双轴图表高级注入 -----------------
                excel_buffer = io.BytesIO()
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    df_excel_clean.to_excel(writer, sheet_name="分时段接通率情况", index=False)
                    st.session_state['df_post_res'].to_excel(writer, sheet_name="上线后明细详单", index=False)
                    if uploaded_pre_file and 'df_pre_res' in st.session_state:
                        st.session_state['df_pre_res'].to_excel(writer, sheet_name="上线前明细详单", index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets["分时段接通率情况"]
                    
                    # 单元格展现样式格式化
                    for row in range(2, 26):
                        for col in [2, 3, 4, 5]:
                            if worksheet.cell(row=row, column=col).value is not None:
                                worksheet.cell(row=row, column=col).number_format = '0.0%'
                        if worksheet.cell(row=row, column=6).value is not None:
                            worksheet.cell(row=row, column=6).number_format = '#,##0.0'
                    
                    # 1. 声明拥有原生 .combine 方法的 LineChart 作为基载轴（左轴）
                    chart_line = LineChart()
                    chart_line.title = "分时段接通率情况与挽回漏接量对比分析"
                    chart_line.style = 13
                    chart_line.y_axis.title = "接通率 (%)"
                    chart_line.y_axis.scaling.min = 0.0
                    chart_line.y_axis.scaling.max = 1.0
                    
                    # 绑定折线图数据（第2列到第4列：3条接通率）
                    data_lines = Reference(worksheet, min_col=2, min_row=1, max_col=4, max_row=25)
                    cats = Reference(worksheet, min_col=1, min_row=2, max_row=25)
                    chart_line.add_data(data_lines, titles_from_data=True)
                    chart_line.set_categories(cats)
                    
                    # 完美开启折线小圆圈标记
                    for series in chart_line.series:
                        series.marker.symbol = "circle"
                        series.marker.size = 5
                    
                    # 2. 声明没有 .combine 方法的 BarChart 作为附载轴（右轴）
                    chart_bar = BarChart()
                    chart_bar.type = "col"
                    
                    # 绑定柱状图数据（第6列：减少电话漏接量）
                    data_bars = Reference(worksheet, min_col=6, min_row=1, max_row=25)
                    chart_bar.add_data(data_bars, titles_from_data=True)
                    
                    # 3. 独立轴解耦跨越配置：将柱状图拉到右侧次轴
                    chart_bar.y_axis.title = "减少电话漏接量 (通)"
                    chart_bar.y_axis.axId = 200
                    chart_bar.y_axis.crosses = "max"          # 柱状图推至右轴
                    chart_line.y_axis.crosses = "autoZero"     # 折线图锁在左轴
                    
                    # 解耦 X 轴冲突：将柱状图的 X 轴完全复用并挂载到折线图的 X 轴通道上
                    chart_bar.x_axis = chart_line.x_axis
                    
                    # 使用标准无冲突接口进行混合注入
                    chart_line.combine(chart_bar)
                    
                    # 将渲染完的高级动态原生图表插入到 A28
                    worksheet.add_chart(chart_line, "A28")
                    
                st.download_button(
                    label="📥 导出带 1:1 原生联动图表的全量 Excel",
                    data=excel_buffer.getvalue(),
                    file_name="南京金鹰世界G酒店_时段复盘全量报告.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("❌ 详单解析失败，请检查基础列名是否完全吻合。")
        except Exception as e:
            st.error(f"分析处理发生非预期异常: {e}")
else:
    st.info("💡 提示：请先在上方数据中心上传分机配置参考表与上线后详单底表。")
